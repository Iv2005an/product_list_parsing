import os
from math import acos, pi

import autocorrect
import openpyxl
import pdf2image
import pytesseract
from PIL import ImageEnhance, Image, ImageOps, ImageDraw


def find_horizontal_line(
        page: Image,
        min_line_length: int = 0,
        max_hole: int = 0,
        y_start: int = 0,
        x_start: int = 0,
        y_finish: int = 0,
        x_finish: int = 0
) -> tuple[tuple[int, int], tuple[int, int], tuple[int, int]] | None:
    print('Finding horizontal line')
    if min_line_length < 1:
        min_line_length = page.width - 1
    if y_finish < 1 or y_finish > page.height:
        y_finish = page.height
    if x_finish < 1 or x_finish > page.width:
        x_finish = page.width
    if x_finish - x_start < min_line_length:
        print('Horizontal line not found')
        return
    for y in range(y_start, y_finish):
        for x in range(x_start, x_finish):
            if page.getpixel((x, y)) == 0:
                x_line_finish = x
                y_line_finish = y_min = y_max = y
                while True:
                    x_line_finish += 1
                    if x_line_finish < x_finish:
                        if page.getpixel((x_line_finish, y_line_finish)) == 0:
                            continue
                    bypassed = False
                    for a_x in range(1, max_hole + 2):
                        if (x_line_finish + a_x < x_finish and
                                page.getpixel((x_line_finish + a_x, y_line_finish)) == 0):
                            x_line_finish += a_x
                            bypassed = True
                            break
                    if bypassed:
                        continue
                    for a_x in range(1, max_hole + 2):
                        for a_y in range(1, max_hole + 2):
                            if (x_line_finish + a_x < x_finish and y_line_finish - a_y > y_start and
                                    page.getpixel((x_line_finish + a_x, y_line_finish - a_y)) == 0):
                                y_line_finish -= a_y
                                y_min = min(y_min, y_line_finish)
                                bypassed = True
                                break
                            if (x_line_finish + a_x < x_finish and y_line_finish + a_y < y_finish and
                                    page.getpixel((x_line_finish + a_x, y_line_finish + a_y)) == 0):
                                y_line_finish += a_y
                                y_max = max(y_max, y_line_finish)
                                bypassed = True
                                break
                        if bypassed:
                            x_line_finish += a_x
                            break
                    if bypassed:
                        continue
                    if x_line_finish - x_start < min_line_length:
                        break
                    x_line_finish -= 1
                    x_line_start = x_line_finish
                    y_line_start = y_line_finish
                    while True:
                        if x_line_start < x_start + 1:
                            break
                        x_line_start -= 1
                        if page.getpixel((x_line_start, y_line_start)) == 0:
                            continue
                        bypassed = False
                        for a_x in range(1, max_hole + 2):
                            if (x_line_start - a_x > x_start and
                                    page.getpixel((x_line_start - a_x, y_line_start)) == 0):
                                x_line_start -= a_x
                                bypassed = True
                                break
                        if bypassed:
                            continue
                        for a_x in range(1, max_hole + 2):
                            for a_y in range(1, max_hole + 2):
                                if (x_line_start - a_x > x_start and y_line_start - a_y > y_start and
                                        page.getpixel((x_line_start - a_x, y_line_start - a_y)) == 0):
                                    y_line_start -= a_y
                                    y_min = min(y_min, y_line_start)
                                    bypassed = True
                                    break
                                if (x_line_start - a_x > x_start and y_line_start + a_y < y_finish and
                                        page.getpixel((x_line_start - a_x, y_line_start + a_y)) == 0):
                                    y_line_start += a_y
                                    y_max = max(y_max, y_line_start)
                                    bypassed = True
                                    break
                            if bypassed:
                                x_line_start -= a_x
                                break
                        if bypassed:
                            continue
                        x_line_start += 1
                        break
                    if x_line_finish - x_line_start >= min_line_length:
                        print('Founded horizontal line:',
                              (x_line_start, y_line_start),
                              (x_line_finish, y_line_finish))
                        return (x_line_start, y_line_start), (x_line_finish, y_line_finish), (y_min, y_max)
                    break
    print('Horizontal line not found')


def find_vertical_line(
        page: Image,
        min_line_length: int = 0,
        max_hole: int = 0,
        x_start: int = 0,
        y_start: int = 0,
        x_finish: int = 0,
        y_finish: int = 0
) -> tuple[tuple[int, int], tuple[int, int], tuple[int, int]] | None:
    print('Finding vertical line')
    if min_line_length < 1:
        min_line_length = page.height - 1
    if x_finish < 1 or x_finish > page.width:
        x_finish = page.width
    if y_finish < 1 or y_finish > page.height:
        y_finish = page.height
    if y_finish - y_start < min_line_length:
        print('Vertical line not found')
        return
    for x in range(x_start, x_finish):
        for y in range(y_start, y_finish):
            if page.getpixel((x, y)) == 0:
                x_line_finish = x_min = x_max = x
                y_line_finish = y
                while True:
                    y_line_finish += 1
                    if y_line_finish < y_finish:
                        if page.getpixel((x_line_finish, y_line_finish)) == 0:
                            continue
                    bypassed = False
                    for a_y in range(1, max_hole + 2):
                        if (y_line_finish + a_y < y_finish and
                                page.getpixel((x_line_finish, y_line_finish + a_y)) == 0):
                            y_line_finish += a_y
                            bypassed = True
                            break
                    if bypassed:
                        continue
                    for a_y in range(1, max_hole + 2):
                        for a_x in range(1, max_hole + 2):
                            if (x_line_finish - a_x > x_start and y_line_finish + a_y < y_finish and
                                    page.getpixel((x_line_finish - a_x, y_line_finish + a_y)) == 0):
                                x_line_finish -= a_x
                                x_min = min(x_min, x_line_finish)
                                bypassed = True
                                break
                            if (x_line_finish + a_x < x_finish and y_line_finish + a_y < y_finish and
                                    page.getpixel((x_line_finish + a_x, y_line_finish + a_y)) == 0):
                                x_line_finish += a_x
                                x_max = max(x_max, x_line_finish)
                                bypassed = True
                                break
                        if bypassed:
                            y_line_finish += a_y
                            break
                    if bypassed:
                        continue
                    if y_line_finish - y_start < min_line_length:
                        break
                    y_line_finish -= 1
                    x_line_start = x_line_finish
                    y_line_start = y_line_finish
                    while True:
                        if y_line_start < y_start + 1:
                            break
                        y_line_start -= 1
                        if page.getpixel((x_line_start, y_line_start)) == 0:
                            continue
                        bypassed = False
                        for a_y in range(1, max_hole + 2):
                            if (y_line_start - a_y > y_start and
                                    page.getpixel((x_line_start, y_line_start - a_y)) == 0):
                                y_line_start -= a_y
                                bypassed = True
                                break
                        if bypassed:
                            continue
                        for a_y in range(1, max_hole + 2):
                            for a_x in range(1, max_hole + 2):
                                if (x_line_start - a_x > x_start and y_line_start - a_y > y_start and
                                        page.getpixel((x_line_start - a_x, y_line_start - a_y)) == 0):
                                    x_line_start -= a_x
                                    x_min = min(x_min, x_line_start)
                                    bypassed = True
                                    break
                                if (x_line_start + a_x < x_finish and y_line_start - a_y > y_start and
                                        page.getpixel((x_line_start + a_x, y_line_start - a_y)) == 0):
                                    x_line_start += a_x
                                    x_max = max(x_max, x_line_start)
                                    bypassed = True
                                    break
                            if bypassed:
                                y_line_start -= a_y
                                break
                        if bypassed:
                            continue
                        y_line_start += 1
                        break
                    if y_line_finish - y_line_start >= min_line_length:
                        print('Founded vertical line:',
                              (x_line_start, y_line_start),
                              (x_line_finish, y_line_finish))
                        return (x_line_start, y_line_start), (x_line_finish, y_line_finish), (x_min, x_max)
                    break
    print('Vertical line not found')


def align_page(page: Image) -> Image:
    print('Aligning page')
    vertical_line = find_vertical_line(
        page,
        int(page.height * 0.75),
        3,
        0,
        0,
        int(page.width * 0.2))
    if vertical_line:
        if vertical_line[0][0] == vertical_line[1][0]:
            return page
        x_a = vertical_line[0][0]
        y_a = vertical_line[0][1]
        x_b = vertical_line[1][0]
        y_b = vertical_line[1][1]
        ab = ((x_b - x_a) ** 2 + (
                y_b - y_a) ** 2) ** 0.5
        ac = ((x_b - x_a) ** 2 + (
                y_a - y_a) ** 2) ** 0.5
        bc = ((x_b - x_b) ** 2 + (
                y_a - y_b) ** 2) ** 0.5
        angle = 90 - acos((ab ** 2 + ac ** 2 - bc ** 2) / (2 * ab * ac)) * (180 / pi)
        if x_a < x_b:
            angle = -angle
        return page.rotate(angle, fillcolor=255, expand=True)


def crop_table(page: Image) -> Image:
    print('Finding top and right sides of table')
    right = 0
    horizontal_line = find_horizontal_line(page, 200, 3)
    for i in range(2):
        if not horizontal_line:
            print('Data table not found')
            return
        cell_width = 0
        left = horizontal_line[0][0] + 5
        first_measure = True
        for right in range(left, int(page.width * 0.95)):
            cell_width = 0
            for y in range(horizontal_line[2][1] + 5, page.height):
                if cell_width > 99 or page.getpixel((right, y)) == 0:
                    break
                cell_width += 1
            if first_measure and (cell_width < 80 or cell_width > 99):
                break
            first_measure = False
            if cell_width > 99:
                noise_count = 0
                for a_x in range(5):
                    cell_width = 0
                    for y in range(horizontal_line[2][1] + 5, page.height):
                        if cell_width > 99 or page.getpixel((right + a_x, y)) == 0:
                            break
                        cell_width += 1
                    if cell_width < 100:
                        noise_count += 1
                if noise_count > 1:
                    continue
                break
        if not first_measure and cell_width > 99:
            break
        if i < 1:
            horizontal_line = find_horizontal_line(page, 200, 3, horizontal_line[2][1] + 10)
    else:
        print('Data table not found')
        return
    right -= 5
    print('Finding left and bottom sides of table')
    left = horizontal_line[0][0] + 5
    top = horizontal_line[2][1] + 5
    cell_height = 0
    last_border_pixel = 0
    for x in range(left, int(page.width * 0.3)):
        if x > last_border_pixel:
            if page.getpixel((x, top)) == 0:
                if 24 < cell_height < 31:
                    left = x - cell_height + 5
                    break
                for a_x in range(1, 5):
                    if page.getpixel((x + a_x, top)) == 0:
                        last_border_pixel = x + a_x
                cell_height = 0
                continue
            last_border_pixel = 0
            cell_height += 1
    else:
        print('Data table not found')
        return
    bottom = top
    for _ in range(10):
        horizontal_line = find_horizontal_line(page, int((right - left) * 0.5), 3, bottom + 30, left, 0, right)
        if not horizontal_line:
            print('Data table not found')
            return
        avg_start_finish_y = horizontal_line[0][1] + abs(horizontal_line[0][1] - horizontal_line[1][1]) // 2
        avg_min_max_y = horizontal_line[2][0] + (horizontal_line[2][1] - horizontal_line[2][0]) // 2
        bottom = min(avg_start_finish_y, avg_min_max_y) + abs(avg_start_finish_y - avg_min_max_y) // 2
    return page.crop((left, top, right, bottom))


def draw_table(page: Image) -> Image:
    horizontal_line = find_horizontal_line(page, page.width - 10, 3)
    vertical_line_min_length = page.height - 10
    first_horizontal_line_y = 0
    if horizontal_line:
        first_horizontal_line_y = horizontal_line[0][1] - abs(horizontal_line[1][1] - horizontal_line[0][1]) // 2
        vertical_line_min_length = page.height - first_horizontal_line_y - 10
    horizontal_lines = []
    while horizontal_line:
        horizontal_lines.append(horizontal_line)
        horizontal_line = find_horizontal_line(page, page.width - 10, 3, horizontal_line[2][1] + 10)
    vertical_line = find_vertical_line(page, vertical_line_min_length, 3)
    page_draw = ImageDraw.Draw(page)
    vertical_lines = []
    while vertical_line:
        vertical_lines.append(vertical_line)
        avg_x = vertical_line[0][0] - abs(vertical_line[1][0] - vertical_line[1][0]) // 2
        top = 0
        if vertical_line[0][1] > 70:
            top = first_horizontal_line_y
        page_draw.line(
            (avg_x, top, avg_x, page.height),
            255, 7)
        vertical_line = find_vertical_line(page, vertical_line_min_length, 3, vertical_line[2][1] + 10)
    for horizontal_line in horizontal_lines:
        avg_y = horizontal_line[0][1] - abs(horizontal_line[1][1] - horizontal_line[0][1]) // 2
        page_draw.line(
            (0, horizontal_line[0][1], page.width, horizontal_line[1][1]),
            255, 7)
        page_draw.line(
            (0, avg_y, page.width, avg_y),
            0, 1)
    for vertical_line in vertical_lines:
        avg_x = vertical_line[0][0] - abs(vertical_line[1][0] - vertical_line[1][0]) // 2
        top = 0
        if vertical_line[0][1] > 70:
            top = first_horizontal_line_y
        page_draw.line(
            (avg_x, top, avg_x, page.height),
            0, 1)
    page_draw.line(
        (page.width - 1, 0, page.width - 1, page.height),
        0, 1)
    page_draw.line(
        (0, page.height - 1, page.width, page.height - 1),
        0, 1)
    return page


def get_cells_positions(page: Image, page_i: int = 0) -> list | None:
    print('Finding cells positions')
    horizontal_line = find_horizontal_line(page)
    if not horizontal_line:
        return
    x0 = 0
    y1 = horizontal_line[0][1]
    vertical_line = find_vertical_line(page, horizontal_line[0][1] - 1, y_finish=y1)
    combined_cells = []
    while vertical_line:
        if vertical_line[0][0] - x0 > 35:
            combined_cells.append(((x0, 0), (vertical_line[0][0], y1)))
        x0 = vertical_line[0][0]
        vertical_line = find_vertical_line(page, horizontal_line[0][1] - 1, x_start=x0 + 1, y_finish=y1)
    vertical_line_min_length = 0
    if len(combined_cells) > 0:
        vertical_line_min_length = page.height - horizontal_line[0][1] - 1
    rows = []
    row_i = 0
    x0 = 0
    vertical_line = find_vertical_line(page, vertical_line_min_length)
    while vertical_line:
        rows.append([])
        x1 = vertical_line[0][0]
        y0 = 0
        horizontal_line_min_length = x1 - x0 - 1
        horizontal_line = find_horizontal_line(page, horizontal_line_min_length,
                                               y_start=70, x_start=x0, x_finish=x1)
        while horizontal_line:
            if y0 == 0:
                for combined_cell in combined_cells:
                    if x0 >= combined_cell[0][0] and x1 <= combined_cell[1][0]:
                        rows[row_i].insert(0, (
                            (combined_cell[0][0], combined_cell[0][1]), (combined_cell[1][0], combined_cell[1][1])))
            rows[row_i].insert(0, ((x0, y0), (x1, horizontal_line[0][1])))
            y0 = horizontal_line[0][1] + 1
            horizontal_line = find_horizontal_line(page, horizontal_line_min_length,
                                                   y_start=y0, x_start=x0, x_finish=x1)
        row_i += 1
        x0 = x1 + 1
        vertical_line = find_vertical_line(
            page, vertical_line_min_length, x_start=x0 + 5)
    if page_i > 0:
        if not os.path.isdir('images/cells'):
            os.mkdir('images/cells')
        if not os.path.isdir(f'images/cells/page_{page_i}'):
            os.mkdir(f'images/cells/page_{page_i}')
        for row_i, row in enumerate(rows):
            if not os.path.isdir(f'images/cells/page_{page_i}/row_{row_i + 1}'):
                os.mkdir(f'images/cells/page_{page_i}/row_{row_i + 1}')
            for cell_i, cell in enumerate(row):
                angle = -90
                if cell_i == len(row) - 1:
                    angle = 180
                page.crop((cell[0][0], cell[0][1], cell[1][0], cell[1][1])).rotate(angle, expand=True).save(
                    f'images/cells/page_{page_i}/row_{row_i + 1}/cell_{cell_i + 1}.png')
    if len(rows) > 0:
        return rows


def get_cells_data(page: Image, cells_positions: list) -> list | None:
    print('Getting cells data')
    speller = autocorrect.Speller('ru')
    cells_data = []
    config = ''
    for row_i, row in enumerate(cells_positions):
        cells_data.append([])
        for cell_i, cell in enumerate(row):
            print(f'Recognizing text ({row_i + 1}, {cell_i + 1})')
            cell_width = cell[1][0] - cell[0][0]
            cell_height = cell[1][1] - cell[0][1]
            cell_max_fill = cell_width * cell_height
            cell_current_fill = 0
            for y in range(cell[0][1], cell[1][1]):
                for x in range(cell[0][0], cell[1][0]):
                    cell_current_fill += page.getpixel((x, y)) / 255
            print(cell_current_fill, cell_max_fill, cell_current_fill / cell_max_fill)
            result = ''
            if cell_current_fill / cell_max_fill < 0.99:
                angle = -90
                if cell_i == len(row) - 1:
                    angle = 180
                cell_image = page.crop((cell[0][0], cell[0][1], cell[1][0], cell[1][1])).rotate(
                    angle, expand=True)
                cell_image = ImageOps.invert(cell_image)
                result = pytesseract.image_to_string(cell_image, lang='rus+eng', config=config)
                result = speller(result)
            cells_data[row_i].append(result)
            print(f'Result: {cells_data[row_i][cell_i]}')
    if len(cells_data) > 0:
        return cells_data


def save_to_excel(cells_data: list, excel_file_path: str, sheet_name: str = None) -> None:
    print('Saving to excel')
    if not os.path.exists(excel_file_path):
        workbook = openpyxl.Workbook()
        if sheet_name:
            workbook.active.title = sheet_name
    else:
        workbook = openpyxl.load_workbook(excel_file_path)
    if sheet_name and sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.create_sheet(sheet_name)
    for row_i, row in enumerate(cells_data):
        for cell_i, cell in enumerate(row):
            if len(cell) > 0:
                worksheet.cell(row_i + 1, cell_i + 1, cell)
    workbook.save(excel_file_path)


def main():
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    if not os.path.isdir('images'):
        os.mkdir('images')
    print('Opening PDF')
    file_name = 'test'
    start_page = 39
    pdf_pages = pdf2image.pdf2image.convert_from_path(f'test_res/test_pdfs/{file_name}.pdf',
                                                      first_page=start_page,
                                                      last_page=start_page,
                                                      dpi=100,
                                                      poppler_path='poppler/Library/bin')
    for page_i, page in enumerate(pdf_pages):
        print(f'Page {start_page + page_i}, working')
        page = ImageEnhance.Contrast(page).enhance(10)
        for y in range(page.height):
            for x in range(page.width):
                pixel = page.getpixel((x, y))
                if max(pixel) - min(pixel) > 20:
                    page.putpixel((x, y), (255, 255, 255))
        page = page.point(lambda p: p > 180 and 255)
        page = page.convert('1')
        page = ImageOps.crop(page, 30)
        page = ImageOps.expand(page, 30, 255)
        page = align_page(page)
        if page:
            page = crop_table(page)
        if page:
            page = draw_table(page)
        if page:
            rows = get_cells_positions(page, start_page + page_i)
            if rows:
                cells_data = get_cells_data(page, rows)
                if cells_data:
                    save_to_excel(cells_data, f'{file_name}.xlsx', f'Страница{start_page + page_i}')
            if not os.path.isdir('images/drawn'):
                os.mkdir('images/drawn')
            page.save(f'images/drawn/{start_page + page_i}.png')
            print(f'Page {start_page + page_i}, complete')


if __name__ == '__main__':
    main()
